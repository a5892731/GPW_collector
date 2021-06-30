#!/bin/python3
'''

author: a5892731
date: 2021-06-29
actualization date: 2021-06-30
version: 1.0

description: This script collects data from Warsaw Stock Exchange (WSE) --- Giełda Papierów Wartościowych w Warszawie (GPW)
             and saves it in Firm_Name.xlsx file in xls_files folder, for each company in WSE.
             If file is not exist then create it and add a data row, else just add row.
source: https://www.money.pl/gielda/spolki-gpw/

'''

import requests
from bs4 import BeautifulSoup
from os import chdir, path
from openpyxl import Workbook, load_workbook
from time import gmtime, strftime

class Data:
    url = 'https://www.money.pl/gielda/spolki-gpw/'
    header_oryginal = ["Walor", "Kurs (PLN)", "Zmiana (%)", "Otwarcie", "Min", "Max",
                       "Obrót (szt.)", "Obrót (PLN)", "Czas_aktualizacji", "Data"]
    header_eng = ["Value", "Rate (PLN)", "Change (%)", "Opening", "Min", "Max",
                  "Turnover (pcs.)", "Turnover (PLN)", "Update_time", "Date"]

    header = header_oryginal

class WSE_Data_Collector(Data):

    def __init__(self):
        r = requests.get(self.url)
        soup = BeautifulSoup(r.text, "html.parser") # "html.parser" is for get rid of warnings > https://stackoverflow.com/questions/33511544/how-to-get-rid-of-beautifulsoup-user-warning
        headline_data = soup.find_all("div") # <div>This is heading</div>
        self.table = self.translate_data(headline_data)

    def translate_data(self, headline_data):
        number_of_headline = 1
        table = []
        row = []
        column_number = 1

        for item in headline_data:  # item is headline with attributes <div>
            if number_of_headline < 121 or item.text == "" or len(item.text) > 40 \
                    or "Loading" in item.text or "MONEY.PL" in item.text\
                    or number_of_headline % 2 == 0 or "——————" in item.text:  # trash filter
                number_of_headline += 1
                continue
            if column_number == 1 and "," in item.text: # trash filter two
                continue
            #print("{}) >>> {}".format(number_of_headline, item.text)) #  this is data test printer

            number_of_headline += 1
            row.append("{}".format(item.text.replace("\xa0","")))

            if len(row) == 9:  # create row in table
                table.append(row)
                #print(row) # this is data test printer two for comtaration
                row = []

            if column_number == 9:
                column_number = 0
            column_number += 1

        return table

class XLS_Creator(Data):

    def __init__(self, row):
        self.row = {}
        self.current_date = strftime("%Y-%m-%d", gmtime())

        for i in range(9):
            self.row[self.header[i]] = row[i]

        self.time_corrector()

        row.append(self.row["{}".format(self.header[9])])
        self.create_xls_file(row)

    def create_xls_file(self, row, title = "Arkusz 1"):

        if "/" in row[0]:
            row[0] = row[0].replace("/", "_")  # id / in company name


        chdir("xls_files")

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

        if path.isfile("{}.xlsx".format(row[0])):
            wb = load_workbook(filename="{}.xlsx".format(row[0]))
            ws = wb.active
            ws.append([column_value for column_value in row])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = title
            for column_nr in range(len(row)):
                ws.column_dimensions[columns[column_nr]].width = 20

            ws.append([column_name for column_name in self.header])
            ws.append([column_value for column_value in row])

        wb.save("{}.xlsx".format(row[0]))
        wb.close()
        chdir("..")

    def time_corrector(self):
        if ":" in self.row["{}".format(self.header[8])]:  # Czas_aktualizacji / Update_time
            self.row["{}".format(self.header[9])] = self.current_date # Data / Date
        else:
            self.row["{}".format(self.header[9])] = self.row["{}".format(self.header[8])]
            self.row["{}".format(self.header[8])] = "—"


#-----------------------------------------------------------------------------------------------------------------------

if __name__ == "__main__":

    GPW = WSE_Data_Collector()
    for row in GPW.table:
        data_storage = XLS_Creator(row)
        print(data_storage.row)

